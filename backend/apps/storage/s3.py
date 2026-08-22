"""
Object-storage access layer.

Everything that talks to S3 goes through here. MinIO and Amazon S3 differ only
in two settings (``endpoint_url`` and path-style addressing), so the rest of the
codebase never needs to know which one is behind it:

    MinIO   -> S3_ENDPOINT_URL=http://minio:9000, S3_PATH_STYLE=1
    AWS S3  -> S3_ENDPOINT_URL unset,             S3_PATH_STYLE=0

Two client flavours are exposed on purpose:

* ``boto3`` for control-plane work (listing, HEAD, small bounded range reads).
  Cheap, no JVM, safe to call from the request/response cycle.
* the ``s3a://`` URI helpers for Spark, which reads bulk data itself so the
  bytes never pass through the web process.
"""
from __future__ import annotations

import csv
import io
import logging
from dataclasses import dataclass, field
from datetime import datetime
from functools import lru_cache
from typing import Any, Iterable

import boto3
from botocore.client import Config
from botocore.exceptions import ClientError
from django.conf import settings

logger = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = (".csv", ".tsv", ".xlsx", ".xls", ".parquet")
EXCEL_EXTENSIONS = (".xlsx", ".xls")

# How much of a CSV we are willing to pull into the web process to infer the
# header and a preview. Bounded on purpose: previewing a 2 GB file must cost
# the same as previewing a 2 KB one.
PREVIEW_BYTE_BUDGET = 256 * 1024


class StorageError(RuntimeError):
    """Raised for anything the caller can sensibly surface to a user."""


class ObjectNotFound(StorageError):
    pass


class ObjectTooLarge(StorageError):
    pass


@dataclass(frozen=True)
class S3Object:
    key: str
    size: int
    last_modified: datetime | None
    extension: str

    @property
    def name(self) -> str:
        return self.key.rsplit("/", 1)[-1]

    def as_dict(self) -> dict[str, Any]:
        return {
            "key": self.key,
            "name": self.name,
            "size": self.size,
            "size_human": human_bytes(self.size),
            "extension": self.extension,
            "last_modified": (
                self.last_modified.isoformat() if self.last_modified else None
            ),
        }


@dataclass
class FilePreview:
    key: str
    columns: list[str] = field(default_factory=list)
    sample_rows: list[dict[str, str]] = field(default_factory=list)
    truncated: bool = True
    sheet_names: list[str] = field(default_factory=list)
    note: str = ""

    def as_dict(self) -> dict[str, Any]:
        return {
            "key": self.key,
            "columns": self.columns,
            "sample_rows": self.sample_rows,
            "truncated": self.truncated,
            "sheet_names": self.sheet_names,
            "note": self.note,
        }


def human_bytes(num: float) -> str:
    for unit in ("B", "KB", "MB", "GB", "TB"):
        if abs(num) < 1024.0:
            return f"{num:3.1f} {unit}".strip()
        num /= 1024.0
    return f"{num:.1f} PB"


@lru_cache(maxsize=1)
def get_client():
    """A process-wide boto3 S3 client (thread-safe for our read-only usage)."""
    return boto3.client(
        "s3",
        endpoint_url=settings.S3_ENDPOINT_URL,
        aws_access_key_id=settings.AWS_ACCESS_KEY_ID or None,
        aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY or None,
        region_name=settings.AWS_DEFAULT_REGION,
        use_ssl=settings.S3_USE_SSL,
        config=Config(
            signature_version="s3v4",
            s3={"addressing_style": "path" if settings.S3_PATH_STYLE else "auto"},
            retries={"max_attempts": 5, "mode": "standard"},
            connect_timeout=10,
            read_timeout=60,
        ),
    )


def s3a_uri(key: str, bucket: str | None = None) -> str:
    """Spark-side URI. Spark reads the bytes; the API never does."""
    return f"s3a://{bucket or settings.S3_BUCKET}/{key.lstrip('/')}"


def extension_of(key: str) -> str:
    _, _, tail = key.rpartition(".")
    return f".{tail.lower()}" if tail and tail != key else ""


def ensure_bucket(bucket: str | None = None) -> None:
    bucket = bucket or settings.S3_BUCKET
    client = get_client()
    try:
        client.head_bucket(Bucket=bucket)
    except ClientError as exc:
        code = exc.response.get("Error", {}).get("Code", "")
        if code not in {"404", "NoSuchBucket", "NotFound"}:
            raise
        logger.info("creating bucket %s", bucket)
        client.create_bucket(Bucket=bucket)


def list_files(
    prefix: str | None = None,
    bucket: str | None = None,
    extensions: Iterable[str] = SUPPORTED_EXTENSIONS,
) -> list[S3Object]:
    """List selectable input files under a prefix, newest first."""
    bucket = bucket or settings.S3_BUCKET
    prefix = settings.S3_RAW_PREFIX if prefix is None else prefix
    wanted = tuple(extensions)
    client = get_client()

    objects: list[S3Object] = []
    try:
        paginator = client.get_paginator("list_objects_v2")
        for page in paginator.paginate(Bucket=bucket, Prefix=prefix):
            for item in page.get("Contents", []):
                key = item["Key"]
                if key.endswith("/"):
                    continue
                ext = extension_of(key)
                if wanted and ext not in wanted:
                    continue
                objects.append(
                    S3Object(
                        key=key,
                        size=int(item.get("Size", 0)),
                        last_modified=item.get("LastModified"),
                        extension=ext,
                    )
                )
    except ClientError as exc:
        raise StorageError(f"cannot list s3://{bucket}/{prefix}: {exc}") from exc

    objects.sort(key=lambda o: (o.last_modified is None, o.last_modified), reverse=True)
    return objects


def head_object(key: str, bucket: str | None = None) -> dict[str, Any]:
    bucket = bucket or settings.S3_BUCKET
    try:
        return get_client().head_object(Bucket=bucket, Key=key)
    except ClientError as exc:
        code = exc.response.get("Error", {}).get("Code", "")
        if code in {"404", "NoSuchKey", "NotFound"}:
            raise ObjectNotFound(f"s3://{bucket}/{key} does not exist") from exc
        raise StorageError(str(exc)) from exc


def read_range(key: str, length: int, bucket: str | None = None) -> bytes:
    """Read at most ``length`` bytes from the start of an object."""
    bucket = bucket or settings.S3_BUCKET
    try:
        response = get_client().get_object(
            Bucket=bucket, Key=key, Range=f"bytes=0-{max(length - 1, 0)}"
        )
        return response["Body"].read()
    except ClientError as exc:
        code = exc.response.get("Error", {}).get("Code", "")
        if code in {"404", "NoSuchKey"}:
            raise ObjectNotFound(f"s3://{bucket}/{key} does not exist") from exc
        raise StorageError(str(exc)) from exc


def download_bytes(key: str, max_bytes: int, bucket: str | None = None) -> bytes:
    """Full download, refused above ``max_bytes`` so the API cannot be OOM'd."""
    meta = head_object(key, bucket=bucket)
    size = int(meta.get("ContentLength", 0))
    if size > max_bytes:
        raise ObjectTooLarge(
            f"{key} is {human_bytes(size)}, above the {human_bytes(max_bytes)} limit"
        )
    buffer = io.BytesIO()
    get_client().download_fileobj(bucket or settings.S3_BUCKET, key, buffer)
    return buffer.getvalue()


def delete_prefix(prefix: str, bucket: str | None = None) -> int:
    """Delete every object under a prefix. Used to clean up job output."""
    bucket = bucket or settings.S3_BUCKET
    client = get_client()
    deleted = 0
    paginator = client.get_paginator("list_objects_v2")
    for page in paginator.paginate(Bucket=bucket, Prefix=prefix):
        batch = [{"Key": item["Key"]} for item in page.get("Contents", [])]
        if not batch:
            continue
        client.delete_objects(Bucket=bucket, Delete={"Objects": batch})
        deleted += len(batch)
    return deleted


# --------------------------------------------------------------------------- #
# preview: bounded, synchronous, safe for the request cycle
# --------------------------------------------------------------------------- #
def _sniff_delimiter(sample: str, key: str) -> str:
    if key.lower().endswith(".tsv"):
        return "\t"
    try:
        return csv.Sniffer().sniff(sample[:8192], delimiters=",;\t|").delimiter
    except csv.Error:
        return ","


def preview_delimited(key: str, rows: int = 5) -> FilePreview:
    raw = read_range(key, PREVIEW_BYTE_BUDGET)
    text = raw.decode("utf-8", errors="replace")
    # Drop a trailing partial line so the CSV reader never sees half a record.
    if "\n" in text:
        text = text[: text.rindex("\n")]
    delimiter = _sniff_delimiter(text, key)
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    try:
        header = next(reader)
    except StopIteration:
        return FilePreview(key=key, note="file appears to be empty")

    columns = [(name or f"col_{i}").strip() for i, name in enumerate(header)]
    sample: list[dict[str, str]] = []
    for row in reader:
        if len(sample) >= rows:
            break
        sample.append({col: (row[i] if i < len(row) else "") for i, col in enumerate(columns)})
    return FilePreview(key=key, columns=columns, sample_rows=sample, truncated=True)


def preview_excel(key: str, rows: int = 5, sheet: str | None = None) -> FilePreview:
    import openpyxl

    payload = download_bytes(key, settings.EXCEL_MAX_BYTES)
    workbook = openpyxl.load_workbook(
        io.BytesIO(payload), read_only=True, data_only=True
    )
    try:
        sheet_names = list(workbook.sheetnames)
        worksheet = workbook[sheet] if sheet else workbook[sheet_names[0]]
        iterator = worksheet.iter_rows(values_only=True)
        try:
            header = next(iterator)
        except StopIteration:
            return FilePreview(key=key, sheet_names=sheet_names, note="sheet is empty")
        columns = [
            str(name).strip() if name is not None else f"col_{i}"
            for i, name in enumerate(header)
        ]
        sample: list[dict[str, str]] = []
        for row in iterator:
            if len(sample) >= rows:
                break
            sample.append(
                {
                    col: ("" if row[i] is None else str(row[i]))
                    for i, col in enumerate(columns)
                    if i < len(row)
                }
            )
        return FilePreview(
            key=key,
            columns=columns,
            sample_rows=sample,
            truncated=True,
            sheet_names=sheet_names,
        )
    finally:
        workbook.close()


def preview_parquet(key: str, rows: int = 5) -> FilePreview:
    import pyarrow.parquet as pq

    payload = download_bytes(key, settings.EXCEL_MAX_BYTES)
    table = pq.read_table(io.BytesIO(payload))
    columns = [c for c in table.column_names if not c.startswith("__")]
    head = table.select(columns).slice(0, rows).to_pylist()
    return FilePreview(
        key=key,
        columns=columns,
        sample_rows=[{k: "" if v is None else str(v) for k, v in row.items()} for row in head],
        truncated=True,
    )


def preview_file(key: str, rows: int = 5, sheet: str | None = None) -> FilePreview:
    """Header + a handful of rows, with a hard cap on bytes read."""
    ext = extension_of(key)
    if ext in EXCEL_EXTENSIONS:
        return preview_excel(key, rows=rows, sheet=sheet)
    if ext == ".parquet":
        return preview_parquet(key, rows=rows)
    if ext in {".csv", ".tsv"}:
        return preview_delimited(key, rows=rows)
    raise StorageError(f"unsupported file type: {ext or key}")
